from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import column_index_from_string
from copy import copy

def get_real_max_row(sheet):
    max_row = sheet.max_row
    for row in range(max_row, 0, -1):
        if any(cell.value is not None and str(cell.value).strip() != "" for cell in sheet[row]):
            return row
    return 0

def copy_columns_with_style(ws_src, ws_dest, src_cols='A:G', src_start_row=1,
                             dest_start_row=1, dest_start_col=1):
    """
    Copy a range of columns from source worksheet to destination worksheet
    including values and styles.

    Args:
        ws_src (Worksheet): Source worksheet.
        ws_dest (Worksheet): Destination worksheet.
        src_cols (str): Column range in A1 notation, e.g., 'A:G'.
        src_start_row (int): Starting row in source worksheet.
        dest_start_row (int): Starting row in destination worksheet.
        dest_start_col (int): Starting column index in destination worksheet.

    Returns:
        int: Number of rows copied.
    """
    col_start_letter, col_end_letter = src_cols.split(":")
    col_start_idx = column_index_from_string(col_start_letter)
    col_end_idx = column_index_from_string(col_end_letter)
    max_row = get_real_max_row(ws_src)

    for i, row in enumerate(range(src_start_row, max_row + 1)):
        for j, col in enumerate(range(col_start_idx, col_end_idx + 1)):
            src_cell = ws_src.cell(row=row, column=col)
            dest_cell = ws_dest.cell(row=dest_start_row + i,
                                     column=dest_start_col + j)

            dest_cell.value = src_cell.value

            if src_cell.has_style:
                dest_cell.font = copy(src_cell.font)
                dest_cell.fill = copy(src_cell.fill)
                dest_cell.border = copy(src_cell.border)
                dest_cell.alignment = copy(src_cell.alignment)
                dest_cell.number_format = copy(src_cell.number_format)
                dest_cell.protection = copy(src_cell.protection)

            if isinstance(dest_cell.value, float):
                dest_cell.number_format = '0.00'

    return max_row - src_start_row

def apply_style(cell, font=None, alignment=None, border=None, fill=None, format=None):
    """
    Apply given styles to a single cell.

    Args:
        cell (Cell): Target cell.
        font (Font, optional): Font style.
        alignment (Alignment, optional): Cell alignment.
        border (Border, optional): Border style.
        fill (PatternFill, optional): Cell background fill.

    Returns:
        None
    """
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if fill:
        cell.fill = fill
    if format:
        cell.number_format = format

def set_style_in_range(ws, cell_range, font=None, alignment=None, border=None, fill=None, format=None):
    """
    Apply styles to a cell range.

    Args:
        ws (Worksheet): Target worksheet.
        cell_range (str): Range in A1 notation (e.g., 'A1:D10').
        font (Font, optional): Font style to apply.
        alignment (Alignment, optional): Alignment to apply.
        border (Border, optional): Border style to apply.
        fill (PatternFill, optional): Background fill to apply.

    Returns:
        None
    """
    if font is None:
        font = Font(color="0000FF", size=9)
    if alignment is None:
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if border is None:
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if fill is None:
        fill = PatternFill(fill_type='solid', start_color='DDEBF7', end_color='DDEBF7')        

    for row in ws[cell_range]:
        for cell in row:
            apply_style(cell, font, alignment, border, fill, format)
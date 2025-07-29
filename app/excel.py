from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import re

import style

def write_excel_header(sheet):
    """
    Write the header of the worksheet with merged cells, titles, styles.

    Args:
        sheet (Worksheet): The target worksheet to write headers to.

    Returns:
        None
    """
    titles = [
        '查詢品號', '展開順序', '階次及子件料號', '本地品名', '自定義欄位一',
        '規格呎吋', '', '', '',
        '累計用量[含損耗]', '鐵板重量/片', '材料費/單價', '鐵板材料費/片',
        '鐵板米數/片', '孔費/片', '折刀/片', '每片小計', '合計'
    ]
    sub_titles = ['材質', '厚度', '長', '寬']

    default_font = Font(color="800000", size=9)
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')

    sheet.merge_cells('F1:I1')
    cell = sheet['F1']
    cell.value = '規格呎吋'
    style.apply_style(cell, default_font, alignment, border, fill)

    for idx, title in enumerate(titles):
        col = idx + 1
        if col in [6, 7, 8, 9]:
            continue
        if title:
            col_letter = get_column_letter(col)
            sheet.merge_cells(f'{col_letter}1:{col_letter}2')
            cell = sheet[f'{col_letter}1']
            cell.value = title
            cell.number_format = '0.00'
            style.apply_style(cell, default_font, alignment, border, fill)

    for i, subtitle in enumerate(sub_titles):
        col_letter = get_column_letter(6 + i)
        cell = sheet[f'{col_letter}2']
        cell.value = subtitle
        style.apply_style(cell, default_font, alignment, border, fill)

    for col in range(1, len(titles) + 1):
        cell = sheet[f'{get_column_letter(col)}2']
        style.apply_style(cell, default_font, alignment, border, fill)

def set_column_widths(sheet):
    """
    Set the width of specified columns.

    :param sheet: openpyxl worksheet
    :param width_map: dict, e.g., {'K': 15, 'L': 12, 'M': 10, 'N': 14}
    """
    width_map = {'C': 12, 'D': 40, 'E': 20}
    for col, width in width_map.items():
        sheet.column_dimensions[col].width = width

def set_basic_styles(sheet, total_row):
    """
    Apply basic styles to sheet.

    Args:
        sheet (Worksheet): Target worksheet.
        total_row: Total row to filled.

    Returns:
        None
    """
    style.set_style_in_range(sheet, f'A3:A{total_row + 3}', alignment=Alignment(horizontal='left', vertical='center', wrap_text=True))
    style.set_style_in_range(sheet, f'B3:B{total_row + 3}', alignment=Alignment(horizontal='right', vertical='center', wrap_text=True))
    style.set_style_in_range(sheet, f'C3:C{total_row + 3}', alignment=Alignment(horizontal='left', vertical='center', wrap_text=True))
    style.set_style_in_range(sheet, f'D3:D{total_row + 3}', alignment=Alignment(horizontal='left', vertical='center', wrap_text=True))
    style.set_style_in_range(sheet, f'E3:E{total_row + 3}', alignment=Alignment(horizontal='left', vertical='center', wrap_text=True))
    style.set_style_in_range(sheet, f'F3:J{total_row + 3}')
    style.set_style_in_range(sheet, f'K3:N{total_row + 3}', fill=PatternFill(fill_type='solid', start_color='FFF2CC', end_color='FFF2CC'))
    style.set_style_in_range(sheet, f'O3:P{total_row + 3}', fill=PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF'))
    style.set_style_in_range(sheet, f'Q3:R{total_row + 3}', fill=PatternFill(fill_type='solid', start_color='F4B084', end_color='F4B084'))
    style.set_style_in_range(sheet, f'J3:J{total_row + 3}', format='0.00')

def extract_parent_and_counts(data):
    """
    Extract parent labels and count the number of associated child items.

    Args:
        data (list): A list of strings from the A column.

    Returns:
        tuple: (parent_labels, child_counts)
    """
    parents = []
    counts = []
    last_parent_index = None

    for i, item in enumerate(data):
        if not item.startswith('。'):  # Parent item
            parents.append(item)
            if last_parent_index is not None:
                counts.append(i - last_parent_index)
            last_parent_index = i

    # Count children for the last parent
    if last_parent_index is not None:
        counts.append(len(data) - last_parent_index)

    return parents, counts

def get_column_content(sheet):
    """
    Retrieve non-empty values from column A, starting from row 5.

    Args:
        sheet (Worksheet): The target worksheet.

    Returns:
        list: List of cell values from column A.
    """
    a_column_values = []
    for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row, min_col=1, max_col=1):
        cell = row[0]
        if cell.value is not None:
            a_column_values.append(cell.value)
    return a_column_values


def get_labels_and_numbers(sheet):
    """
    Get parent labels and corresponding counts from column A.

    Args:
        sheet (Worksheet): The target worksheet.

    Returns:
        tuple: (labels, label_counts)
    """
    values = get_column_content(sheet)
    return extract_parent_and_counts(values)


def fill_query_no(sheet, labels, label_nums):
    """
    Fill 'Query Item No.' values in column A starting from row 3
    based on parent labels and their respective counts.

    Args:
        sheet (Worksheet): The target worksheet.
        labels (list): List of parent labels.
        label_nums (list): List of child counts for each label.

    Returns:
        None
    """
    row = 3
    for label, count in zip(labels, label_nums):
        for _ in range(count):
            sheet[f'A{row}'] = label
            sheet[f'B{row}'] = row - 2
            row += 1

def get_row_range_values(sheet, row_num, start_col, end_col):
    """
    Retrieve values from a specific row and a range of columns.

    :param sheet: An openpyxl worksheet object
    :param row_num: The row number to access (1 = first row)
    :param start_col: Starting column number (1 = column A)
    :param end_col: Ending column number (inclusive)
    :return: List of values from start_col to end_col in the specified row
    """
    return [
        sheet.cell(row=row_num, column=col).value
        for col in range(start_col, end_col + 1)
    ]

def find_row_with_multiple_conditions(sheet, conditions, return_cols=None, start_row=2):
    """
    Search for the first row matching multiple column-value conditions,
    and return values from specified columns in that row.

    :param sheet: openpyxl worksheet object
    :param conditions: list of (col_number, expected_value) tuples
                       e.g. [(1, 'ABC123'), (3, 50)]
    :param return_cols: list of column numbers to return values from
                        e.g. [2, 4]. If None, return the entire row as values.
    :param start_row: row number to start searching from (default: 2 to skip header)
    :return: list of cell values from the matched row (specified columns only),
             or None if no match is found
    """
    for row in sheet.iter_rows(min_row=start_row):
        if all(row[col - 1].value == expected for col, expected in conditions):
            if return_cols:
                return [row[col - 1].value for col in return_cols]
            else:
                return [cell.value for cell in row]  # return full row values
    return None


def find_value_by_match(sheet, match_col, match_value, return_col, start_row=2):
    """
    Search for a specific value in a given column and return the value from another column in the same row.

    :param sheet: An openpyxl worksheet object
    :param match_col: The column to search in (1 = column A)
    :param match_value: The value to look for
    :param return_col: The column to return the value from (1 = column A)
    :param start_row: The row to start searching from (default is 2, assuming row 1 is the header)
    :return: The matched row's value from the return_col, or None if not found
    """
    for row in sheet.iter_rows(min_row=start_row):
        if row[match_col - 1].value == match_value:
            return row[return_col - 1].value
    return None

def parse_thickness(thickness_str):
    if thickness_str is None:
        return None
    match = re.match(r"([\d\.]+)", thickness_str)
    if match:
        return float(match.group(1))
    return None

def find_value_by_thickness(sheet, material, input_thickness):
    candidates = []
    for row in sheet.iter_rows(min_row=2):
        cell_material = row[0].value
        cell_thickness_str = row[1].value
        cell_thickness = parse_thickness(cell_thickness_str)

        if cell_material == material and cell_thickness is not None:
            if cell_thickness >= input_thickness:
                candidates.append((cell_thickness, row[5].value))
    if not candidates:
        return None

    candidates.sort(key=lambda x: x[0])
    return candidates[0][1]

def calculate_and_write_output(base_sheet, output_sheet, weight_sheet, material_sheet, mm_sheet,
                               total_row, base_row, output_row):
    output_col = 11

    for i in range(base_row, base_row + total_row):
        values = get_row_range_values(base_sheet, i, 4, 7)
        output_i = output_row + (i - base_row)

        if None in values:
            for j in range(4):
                output_sheet.cell(row=output_i, column=output_col + j, value="")
            cell1 = output_sheet.cell(row=output_i, column=output_col + 6, value=f'=P{output_i}+O{output_i}+N{output_i}+M{output_i}')
            cell2 = output_sheet.cell(row=output_i, column=output_col + 7, value=f'=Q{output_i}*J{output_i}')
            cell1.number_format = '0.00'
            cell2.number_format = '0.00'
            continue
        
        try:
            # resolve the type 'str'
            for idx in range(1, 4):
                if isinstance(values[idx], str):
                    if '.' in values[idx]:
                        values[idx] = float(values[idx])
                    else:
                        values[idx] = int(values[idx])

            coefficient = float(find_value_by_match(weight_sheet, 2, values[0], 6))
            weight = values[1] * values[2] * values[3] * coefficient
            material = find_row_with_multiple_conditions(material_sheet, [(3, values[0]), (4, values[1])], [5])[0]
            iron = weight * material
            mm = find_value_by_thickness(mm_sheet, values[0], values[1])
            iron_mm = (values[2] + values[3]) * mm * 2 / 1000

            res = [round(weight, 2), round(material, 2), round(iron, 2), round(iron_mm, 2)]

            for j, val in enumerate(res):
                cell = output_sheet.cell(row=output_i, column=output_col + j, value=val)
                cell.number_format = '0.00'

        except Exception as e:
            for j in range(4):
                output_sheet.cell(row=output_i, column=output_col + j, value="")
            error_message = f"Row {i} error: {e}, values: {values}"
            print(error_message)
            raise Exception(error_message)
        
        cell1 = output_sheet.cell(row=output_i, column=output_col + 6, value=f'=P{output_i}+O{output_i}+N{output_i}+M{output_i}')
        cell2 = output_sheet.cell(row=output_i, column=output_col + 7, value=f'=Q{output_i}*J{output_i}')
        cell1.number_format = '0.00'
        cell2.number_format = '0.00'

def total_result(sheet, total_row):
    output_col = 17
    sheet.merge_cells(f'P{total_row + 4}:Q{total_row + 4}')
    sheet.merge_cells(f'P{total_row + 5}:Q{total_row + 5}')
    cell_1 = sheet[f'P{total_row + 4}']
    cell_2 = sheet[f'P{total_row + 5}']
    cell_1.value = '合計'
    cell_2.value = '含其它製程費用'
    style.set_style_in_range(sheet, f'P{total_row + 4}:Q{total_row + 5}', fill=PatternFill(fill_type='solid', start_color='F4B084', end_color='F4B084'))
    cell1 = sheet.cell(row=total_row + 4, column=output_col + 1, value=f'=SUM(R4:R{total_row + 3})')
    cell2 = sheet.cell(row=total_row + 5, column=output_col + 1, value=f'=R{total_row + 4}*1.05*1.3')
    cell1.number_format = '0.00'
    cell2.number_format = '0.00'
    style.set_style_in_range(sheet, f'Q{total_row + 4}:R{total_row + 5}', fill=PatternFill(fill_type='solid', start_color='F4B084', end_color='F4B084'))